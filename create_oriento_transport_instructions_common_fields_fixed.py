import json
import os
import shutil
from openpyxl import load_workbook

JSON_FILE = 'ORIENTO-common-fields-3-template-3-fixed.json'
TRANSPORT_TEMPLATE = 'TEMPLATE-TRANSPORT-INSTRUCTION-BAXXX-CAT-XXX-TRANSPORTER-2-3.xlsx'
OUTPUT_DIR = 'generated_oriento_transport_instructions'


def write_cell(ws, cell_ref, value):
    if cell_ref and value not in (None, ''):
        ws[cell_ref] = value


def expand_range_vertical(cell_range):
    start, end = cell_range.split(':')
    col = ''.join([c for c in start if c.isalpha()])
    row1 = int(''.join([c for c in start if c.isdigit()]))
    row2 = int(''.join([c for c in end if c.isdigit()]))
    return [f'{col}{r}' for r in range(row1, row2 + 1)]


def merged_sheet_fields(common_fields, overrides):
    out = dict(common_fields or {})
    out.update(overrides or {})
    return out


def main():
    with open(JSON_FILE, 'r', encoding='utf-8') as f:
        config = json.load(f)

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    common_fields = config.get('common_sheet_fields', {})

    for instruction in config['instructions']:
        copy_no = instruction['copy_no']
        fields = merged_sheet_fields(common_fields, instruction.get('sheet_field_overrides', {}))
        ref_nr = fields.get('A46_ref_nr', '') or f'COPY_{copy_no:02d}'
        out_file = os.path.join(OUTPUT_DIR, f'ORIENTO_TI_{copy_no:02d}_{ref_nr}.xlsx')
        shutil.copyfile(TRANSPORT_TEMPLATE, out_file)

        wb = load_workbook(out_file)
        ws = wb[wb.sheetnames[0]]

        for key, value in fields.items():
            cell_ref = key.split('_', 1)[0]
            write_cell(ws, cell_ref, value)

        cargo = instruction.get('loading_cargo_details', {})
        target_cells = expand_range_vertical(cargo.get('range', 'F43:F50'))
        rows = cargo.get('rows', [])
        for idx, row in enumerate(rows):
            if idx >= len(target_cells):
                break
            description = row.get('description', '')
            measurement = row.get('measurement', '')
            gross_weight = row.get('gross_weight_kg', '')
            parts = [p for p in [description, measurement, str(gross_weight) if gross_weight not in ('', None) else ''] if p]
            write_cell(ws, target_cells[idx], ' | '.join(parts))

        wb.save(out_file)


if __name__ == '__main__':
    main()
