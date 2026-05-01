import json
import os
import shutil
from openpyxl import load_workbook

JSON_FILE = 'transport_instruction_excel_ordered_template.json'
TRANSPORT_TEMPLATE = '2604DSI2801-TRANSPORT-INSTRUCTION-BA3195-CAT-6030-NATRANS-LOAD1-2.xlsx'
OUTPUT_DIR = 'generated_transport_instructions'


def write_cell(ws, cell_ref, value):
    if cell_ref and value not in (None, ''):
        ws[cell_ref] = value


def expand_range_vertical(cell_range):
    start, end = cell_range.split(':')
    col = ''.join([c for c in start if c.isalpha()])
    row1 = int(''.join([c for c in start if c.isdigit()]))
    row2 = int(''.join([c for c in end if c.isdigit()]))
    return [f'{col}{r}' for r in range(row1, row2 + 1)]


def main():
    with open(JSON_FILE, 'r', encoding='utf-8') as f:
        config = json.load(f)

    os.makedirs(OUTPUT_DIR, exist_ok=True)

    for instruction in config['instructions']:
        copy_no = instruction['copy_no']
        out_file = os.path.join(OUTPUT_DIR, f'TI_copy_{copy_no:02d}.xlsx')
        shutil.copyfile(TRANSPORT_TEMPLATE, out_file)

        wb = load_workbook(out_file)
        ws = wb[wb.sheetnames[0]]

        # Write fixed sheet fields where the key starts with the cell address
        for key, value in instruction.get('sheet_fields', {}).items():
            cell_ref = key.split('_', 1)[0]
            write_cell(ws, cell_ref, value)

        # Write cargo descriptions into F43:F50
        cargo = instruction.get('loading_cargo_details', {})
        target_cells = expand_range_vertical(cargo.get('range', 'F43:F50'))
        rows = cargo.get('rows', [])

        for idx, row in enumerate(rows):
            if idx >= len(target_cells):
                break
            description = row.get('description', '')
            measurement = row.get('measurement', '')
            gross_weight = row.get('gross_weight_kg', '')

            parts = [p for p in [description, measurement, str(gross_weight) if gross_weight not in ('', None) else ''] if p]
            ws[target_cells[idx]] = ' | '.join(parts)

        wb.save(out_file)

if __name__ == '__main__':
    main()